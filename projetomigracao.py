from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
#biblioteca pra gerar planilha
import openpyxl
#biblioteca pra tempo
import math
#biblioteca de tempo de espera antes da ação 
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

#######
# Solicitar ao usuário que insira o caminho da planilha
planilha_gerada = input("Insira o nome que deja salvar o novo arquivo. Não esqueça de colocar '.xlsx' no final : ")
# Solicitar ao usuário que insira o caminho da planilha
caminho_planilha = input("Insira o caminho da planilha que irá puxar os dados: ")

# Carregar a planilha usando o caminho fornecido pelo usuário
try:
    workbook = openpyxl.load_workbook(caminho_planilha)
    
    # Solicitar ao usuário que insira o nome da planilha desejada
    nome_planilha = input("Insira o nome da ABA da planilha: ")
    
    # Verificar se o nome da planilha fornecido pelo usuário existe no workbook
    if nome_planilha in workbook.sheetnames:
        sheet = workbook[nome_planilha]
        # Faça algo com a sheet carregada...
        print("Planilha carregada com sucesso!")
    else:
        print("A planilha com o nome fornecido não existe.")
except FileNotFoundError:
    print("Arquivo não encontrado. Verifique o caminho e tente novamente.")
except Exception as e:
    print("Ocorreu um erro ao carregar a planilha:", str(e))

# Criar uma lista para armazenar os nomes
nomes = []

# Ler os nomes da planilha e adicionar na lista
for row in sheet.iter_rows(min_row=2, values_only=True):  #começa na segunda linha para ignorar o cabeçalho
    nome = row[0]  #row índice da coluna que contém os dados
    nomes.append(nome)
    print(nome)

servico = Service(ChromeDriverManager().install())
navegador = webdriver.Chrome(service=servico)

# abrir o site desejado
navegador.get("https://colegiozerohum.sistematutor.com.br/.......etc")
# preencher campo de login
campo_login = navegador.find_element(By.XPATH, '//*[@id="login"]')
campo_login.send_keys("coloque aqui o e-mail para login")

# preencher campo de senha
campo_senha = navegador.find_element(By.XPATH, '/html/body/section/div/div[2]/div/form/div[2]/input')
campo_senha.send_keys("coloque aqui a senha do e-mail para login")

# enter no login
campo_senha.submit()

#Passo a passo dentro do tutor

#Passo 0, clicar no botão para abrir menu com opção de pesquisar

botao = navegador.find_element(By.XPATH, '//*[@id="toggleSidebarPrimary"]')
botao.click()
#tempo em segundos em que a pagina irá aguardar até a próxima função 
#time.sleep(1)

#Passo 1, pesquisa dentro do Tutor sendo necessário primeiramente definir um nome 
campo_pesquisa_aluno = navegador.find_element(By.XPATH, '//*[@id="pageWrapperPrimary"]/div/div[1]/section/div[2]/div[1]/div[2]/form/div/div[1]/input')
campo_pesquisa_aluno.send_keys("NOME QUALQUER")
#time.sleep(1)

#Passo 2, clicar no botão após definir o nome
wait = WebDriverWait(navegador, 10)
elementa = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pageWrapperPrimary"]/div/div[1]/section/div[2]/div[1]/div[2]/form/div/div[2]/div/button/i')))
#botao_pesquisa = navegador.find_element(By.XPATH, '//*[@id="pageWrapperPrimary"]/div/div[1]/section/div[2]/div[1]/div[2]/form/div/div[2]/div/button/i')
elementa.click()
#time.sleep(1)

#Passo 3, clicar no alterar 
botao_alterar_nome = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ListaAluno"]/tbody/tr/td[7]/button')))
#botao_alterar_nome = navegador.find_element(By.XPATH, '//*[@id="ListaAluno"]/tbody/tr/td[7]/button')
botao_alterar_nome.click()
#time.sleep(1)



#encontrar o elemento do corpo da página
body_element = navegador.find_element(By.TAG_NAME, 'body')

#rolar até o final da página
body_element.send_keys(Keys.END)

#time.sleep(3)

 # Criar um novo arquivo .xlsx
workbook = openpyxl.Workbook()
sheet = workbook.active


total_nomes = len(nomes)  # Total de nomes na lista


for index, nome in enumerate(nomes, start=1):
     
        #aguardar até que o elemento esteja clicável
        wait = WebDriverWait(navegador, 10)
        elemento = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="basico"]/div[16]/div[2]/div/span/button[2]')))

        #clicar no elemento
        elemento.click()

        #Essa é a frame em que o elemento ta dentro 
        frame = navegador.find_element(By.XPATH, '//*[@id="modalFrame2"]')
        navegador.switch_to.frame(frame)

        #Elemento procurado dentro do frame
        campo_pesquisa_financeiro = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[1]/div/div/input')
        campo_pesquisa_financeiro.send_keys(nome) 

        #click em pesquisar dentro do nome do financeiro
        botao_pesquisa_financeiro = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[2]/button')
        botao_pesquisa_financeiro.click()

        #click em pesquisar dentro do nome do financeiro
        botao_pesquisar_financeiro = navegador.find_element(By.XPATH, '//*[@id="ListaAluno"]/tbody/tr/td[4]/input[1]')
        botao_pesquisar_financeiro.click()

       
        #Buscar os dados desejado
        


        #puxar e printar cidade
         
        try:
            cidade_resp_finac = navegador.find_element(By.XPATH,'/html/body/div/form/section/div[1]/div[4]/div[2]/select')
            opcao_selecionada = cidade_resp_finac.find_element(By.CSS_SELECTOR, 'option[selected="selected"]')
            informacao_selecionada1 = opcao_selecionada.text.strip()
            #print(informacao_selecionada1)
        except NoSuchElementException:
            print("Elemento não encontrado.")
              #puxar e printar bairro
        
        try:
            bairro_resp_finac = navegador.find_element(By.XPATH,'/html/body/div/form/section/div[1]/div[4]/div[3]')
            opcao_selecionada = bairro_resp_finac.find_element(By.CSS_SELECTOR, 'option[selected="selected"]')
            informacao_selecionada2 = opcao_selecionada.text.strip()
            #print(informacao_selecionada2)
        except NoSuchElementException:
            print("Elemento não encontrado2.")  

        #puxar e printar UF
        
        try:
            uf_resp_finac = navegador.find_element(By.XPATH,'/html/body/div/form/section/div[1]/div[4]/div[1]')
            opcao_selecionada = uf_resp_finac.find_element(By.CSS_SELECTOR, 'option[selected="selected"]')
            informacao_selecionada3 = opcao_selecionada.text.strip()
            #print(informacao_selecionada3)
        except NoSuchElementException:
            print("Elemento não encontrado3.")  


        #puxar e printar Nome
        nome_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[1]/div[2]/div/input')
        #print(nome_resp_finac.get_attribute("value"))


        #puxar e printar CPF
        cpf_resp_finac = navegador.find_element(By.XPATH, '//*[@id="cpf"]')
        #print(cpf_resp_finac.get_attribute("value"))


        #puxar e printar Email
        email_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[1]/div[3]/div/input')
        #print(email_resp_finac.get_attribute("value"))


        #puxar e printar RG
        rg_resp_finac = navegador.find_element(By.XPATH, '//*[@id="divPessoaFisica1"]/div[1]/input')
        #print(rg_resp_finac.get_attribute("value"))


        #puxar e printar Telefone
        telefone_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[6]/div[2]/div/input')
        #print(telefone_resp_finac.get_attribute("value"))


        #puxar e printar Data de Nascimento
        dtnasci_resp_finac = navegador.find_element(By.XPATH, '//*[@id="dataFormNascimento"]')
        #print(dtnasci_resp_finac.get_attribute("value"))


        #puxar e printar Sexo
        sexo_resp_finac = navegador.find_element(By.XPATH, '//*[@id="divPessoaFisica2"]/div[1]/select')
        #print(sexo_resp_finac.get_attribute("value"))


        #puxar e printar Logradouro 
        logradouro_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[3]/div[2]/div/input')
        #print(logradouro_resp_finac.get_attribute("value"))


        #puxar e printar Complemento
        complemento_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[3]/div[4]/input')
        #print(complemento_resp_finac.get_attribute("value"))


        #puxar e printar CEP
        cep_resp_finac = navegador.find_element(By.XPATH, '//*[@id="cep"]')
        #print(cep_resp_finac.get_attribute("value"))

        #time.sleep(3)


        # Criar um novo arquivo .xlsx
        #workbook = openpyxl.Workbook()
        #sheet = workbook.active

        # Dados a serem enviados
        nome_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[1]/div[2]/div/input').get_attribute("value")
        cpf_resp_finac = navegador.find_element(By.XPATH, '//*[@id="cpf"]').get_attribute("value")
        email_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[1]/div[3]/div/input').get_attribute("value")
        rg_resp_finac = navegador.find_element(By.XPATH, '//*[@id="divPessoaFisica1"]/div[1]/input').get_attribute("value")
        telefone_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[6]/div[2]/div/input').get_attribute("value")
        dtnasci_resp_finac = navegador.find_element(By.XPATH, '//*[@id="dataFormNascimento"]').get_attribute("value")
        sexo_resp_finac = navegador.find_element(By.XPATH, '//*[@id="divPessoaFisica2"]/div[1]/select').get_attribute("value")
        logradouro_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[3]/div[2]/div/input').get_attribute("value")
        complemento_resp_finac = navegador.find_element(By.XPATH, '/html/body/div/form/section/div[1]/div[3]/div[4]/input').get_attribute("value")
        cep_resp_finac = navegador.find_element(By.XPATH, '//*[@id="cep"]').get_attribute("value")
        cidade_resp_finac = informacao_selecionada1
        bairro_resp_finac = informacao_selecionada2
        uf_resp_finac  = informacao_selecionada3

        # Insere os dados na planilha
        dados = [nome_resp_finac,cpf_resp_finac,email_resp_finac, rg_resp_finac, telefone_resp_finac, dtnasci_resp_finac, sexo_resp_finac, logradouro_resp_finac, complemento_resp_finac, cep_resp_finac, bairro_resp_finac, cidade_resp_finac, uf_resp_finac]
        sheet.append(dados)

        #fechar o frame
        navegador.switch_to.default_content()

        #voltar na pesquisa
        botao_fechar_pesquisa = navegador.find_element(By.XPATH, '//*[@id="modal2"]/div/div/div[1]/button/span')
        botao_fechar_pesquisa.click()
        #time.sleep(3)

        
        # Salva o arquivo após cada nome processado
        workbook.save(planilha_gerada)

    
        nomes_restantes = total_nomes - index
        tempo_estimado = nomes_restantes * 4  # Tempo estimado em segundos
        minutos_estimados = math.ceil(tempo_estimado / 60)  # Tempo estimado em minutos (arredondado para cima)

        print(f"*****************************************************************************************************")
        print(f"*Aguarde um pouco, a Dudinha está trabalhando.")
        print(f"*Enquanto isso, pode ir no banheiro. Te aviso quando eu terminar, acompanhe meu progresso abaixo:")
        print(f"*CPF processado: {nome}.  ")
        print(f"*Faltam {nomes_restantes} nomes.")
        print(f"*Tempo estimado: {minutos_estimados} minutos.")
        print(f"*****************************************************************************************************")



# Salva o arquivo
workbook.save(planilha_gerada)

print('Planilha .xlsx gerada com sucesso!')

